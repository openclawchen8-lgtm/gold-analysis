"""
Report Service - 報告生成系統
支援日報/周報/月報自動生成，PDF + Excel 導出
"""
import logging
import io
import json
from typing import Optional
from datetime import datetime, timedelta, date
from dataclasses import dataclass, asdict
from enum import Enum

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.decision import Decision, DecisionType, DecisionSource
from app.models.portfolio import Portfolio
from app.models.portfolio_holding import PortfolioHolding
from app.models.alert import Alert
from app.services.backtest_service import BacktestService
from app.analysis.performance import PerformanceAnalyzer

logger = logging.getLogger(__name__)


class ReportType(str, Enum):
    """報告類型"""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"


@dataclass
class ReportData:
    """報告數據容器"""
    report_type: str
    period_start: datetime
    period_end: datetime
    title: str
    summary: dict
    market_overview: dict
    decisions: list[dict]
    portfolio_performance: dict
    alerts: dict
    generated_at: datetime


class ReportService:
    """報告生成服務"""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.backtest_service = BacktestService(session)

    # ── 數據聚合 ─────────────────────────────────────────────────────────────

    async def _get_decisions_for_period(
        self,
        user_id: int,
        start: datetime,
        end: datetime,
    ) -> list[Decision]:
        stmt = (
            select(Decision)
            .where(
                Decision.user_id == user_id,
                Decision.created_at >= start,
                Decision.created_at <= end,
            )
            .order_by(Decision.created_at.desc())
        )
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_portfolios(self, user_id: int) -> list[Portfolio]:
        stmt = select(Portfolio).where(Portfolio.user_id == user_id)
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_alerts_for_period(
        self,
        user_id: int,
        start: datetime,
        end: datetime,
    ) -> list[Alert]:
        stmt = (
            select(Alert)
            .where(
                Alert.user_id == user_id,
                Alert.created_at >= start,
                Alert.created_at <= end,
            )
            .order_by(Alert.created_at.desc())
        )
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def collect_report_data(
        self,
        user_id: int,
        report_type: ReportType,
        reference_date: Optional[datetime] = None,
    ) -> ReportData:
        """
        收集報告所需數據

        Args:
            user_id: 用戶ID
            report_type: 報告類型
            reference_date: 參考日期（預設今天）

        Returns:
            ReportData
        """
        ref = reference_date or datetime.utcnow()

        if report_type == ReportType.DAILY:
            start = ref.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(days=1) - timedelta(seconds=1)
        elif report_type == ReportType.WEEKLY:
            # 本週一
            start = ref - timedelta(days=ref.weekday())
            start = start.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(days=7) - timedelta(seconds=1)
        else:  # MONTHLY
            start = ref.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            # 下月第一天
            if ref.month == 12:
                next_month = ref.replace(year=ref.year + 1, month=1, day=1)
            else:
                next_month = ref.replace(month=ref.month + 1, day=1)
            end = next_month - timedelta(seconds=1)

        decisions = await self._get_decisions_for_period(user_id, start, end)
        portfolios = await self._get_portfolios(user_id)
        alerts = await self._get_alerts_for_period(user_id, start, end)

        # 市場概覽（存根）
        market_overview = await self._market_overview()

        # 決策統計
        decision_stats = self._summarize_decisions(decisions)

        # 組合績效
        portfolio_perf = {}
        for p in portfolios:
            try:
                perf = await self._calc_portfolio_perf(p)
                portfolio_perf[p.name] = perf
            except Exception as e:
                logger.warning(f"Failed to calc perf for portfolio {p.id}: {e}")

        # 告警統計
        alert_stats = self._summarize_alerts(alerts)

        title_map = {
            ReportType.DAILY: f"黃金分析日報 {ref.strftime('%Y-%m-%d')}",
            ReportType.WEEKLY: f"黃金分析週報 {ref.strftime('%Y-W%U')}",
            ReportType.MONTHLY: f"黃金分析月報 {ref.strftime('%Y年%m月')}",
        }

        return ReportData(
            report_type=report_type.value,
            period_start=start,
            period_end=end,
            title=title_map[report_type],
            summary={
                "decision_count": len(decisions),
                "portfolio_count": len(portfolios),
                "alert_count": len(alerts),
                "triggered_alerts": sum(1 for a in alerts if a.triggered_at),
            },
            market_overview=market_overview,
            decisions=decision_stats,
            portfolio_performance=portfolio_perf,
            alerts=alert_stats,
            generated_at=datetime.utcnow(),
        )

    def _summarize_decisions(self, decisions: list[Decision]) -> list[dict]:
        """決策統計"""
        by_type: dict[str, int] = {}
        by_source: dict[str, int] = {}
        details = []

        for d in decisions:
            t = d.decision_type.value
            s = d.source.value
            by_type[t] = by_type.get(t, 0) + 1
            by_source[s] = by_source.get(s, 0) + 1

            details.append({
                "id": d.id,
                "type": t,
                "asset": d.asset,
                "signal_strength": d.signal_strength,
                "confidence": d.confidence,
                "is_executed": d.is_executed,
                "created_at": d.created_at.isoformat(),
                "reason_zh": d.reason_zh,
            })

        return {
            "total": len(decisions),
            "by_type": by_type,
            "by_source": by_source,
            "details": details,
        }

    def _summarize_alerts(self, alerts: list[Alert]) -> dict:
        """告警統計"""
        total = len(alerts)
        triggered = sum(1 for a in alerts if a.triggered_at)
        active = sum(1 for a in alerts if a.is_active)
        by_type = {}
        for a in alerts:
            by_type[a.alert_type.value] = by_type.get(a.alert_type.value, 0) + 1

        return {
            "total": total,
            "triggered": triggered,
            "active": active,
            "by_type": by_type,
        }

    async def _market_overview(self) -> dict:
        """市場概覽（存根，未來可接入真實行情）"""
        return {
            "gold_price": 2000.0,
            "dxy_index": 104.5,
            "fed_rate": 5.25,
            "note": "Market data stub — replace with live feed",
        }

    async def _calc_portfolio_perf(self, portfolio: Portfolio) -> dict:
        """計算組合績效"""
        holdings = portfolio.holdings or []
        invested = sum(h.avg_cost * h.quantity for h in holdings)
        market_value = sum((h.current_price or h.avg_cost) * h.quantity for h in holdings)
        unrealized = market_value - invested
        unrealized_pct = (unrealized / invested * 100) if invested > 0 else 0.0

        return {
            "initial_capital": portfolio.initial_capital,
            "current_value": portfolio.current_value,
            "invested": round(invested, 2),
            "market_value": round(market_value, 2),
            "unrealized_pnl": round(unrealized, 2),
            "unrealized_pnl_pct": round(unrealized_pct, 2),
            "position_count": len(holdings),
        }

    # ── Markdown 報告 ───────────────────────────────────────────────────────

    def generate_markdown(self, data: ReportData) -> str:
        """生成 Markdown 格式報告"""
        lines = [
            f"# {data.title}",
            "",
            f"**生成時間**: {data.generated_at.strftime('%Y-%m-%d %H:%M:%S')} UTC",
            f"**報告週期**: {data.period_start.strftime('%Y-%m-%d')} ~ {data.period_end.strftime('%Y-%m-%d')}",
            "",
            "---",
            "",
            "## 📊 市場概覽",
            "",
            f"| 品種 | 數值 |",
            f"|------|------|",
        ]

        for key, val in data.market_overview.items():
            if key != "note":
                lines.append(f"| {key} | {val} |")

        lines.extend(["", "## 📋 本期決策回顧", ""])
        ds = data.decisions
        lines.append(f"本期共產生 **{ds.get('total', 0)}** 個決策")
        lines.append("")
        for t, cnt in ds.get("by_type", {}).items():
            lines.append(f"- **{t}**: {cnt}")
        lines.append("")

        if data.portfolio_performance:
            lines.extend(["", "## 💼 組合績效", ""])
            for name, perf in data.portfolio_performance.items():
                lines.append(f"### {name}")
                lines.append(f"- 初始資金: ${perf['initial_capital']:,.2f}")
                lines.append(f"- 當前市值: ${perf['current_value']:,.2f}")
                lines.append(f"- 未實現損益: ${perf['unrealized_pnl']:,.2f} ({perf['unrealized_pnl_pct']:+.2f}%)")
                lines.append("")

        if data.alerts.get("total", 0) > 0:
            lines.extend(["", "## 🔔 告警統計", ""])
            lines.append(f"- 總告警: {data.alerts['total']}")
            lines.append(f"- 已觸發: {data.alerts['triggered']}")
            lines.append(f"- 活躍: {data.alerts['active']}")

        lines.extend(["", "---", f"*報告自動生成於 {datetime.utcnow().isoformat()}*"])
        return "\n".join(lines)

    # ── PDF 導出 ─────────────────────────────────────────────────────────────

    async def generate_pdf(
        self,
        user_id: int,
        report_type: ReportType,
        reference_date: Optional[datetime] = None,
    ) -> bytes:
        """
        生成 PDF 報告

        需要 reportlab: pip install reportlab
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
            from reportlab.lib import colors
        except ImportError:
            logger.error("reportlab not installed, cannot generate PDF")
            return b"ERROR: reportlab not installed"

        data = await self.collect_report_data(user_id, report_type, reference_date)
        md = self.generate_markdown(data)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story = []

        # 解析 Markdown 簡單內容
        for line in md.split("\n"):
            line = line.strip()
            if not line or line.startswith("---"):
                story.append(Spacer(1, 0.3*cm))
            elif line.startswith("# "):
                story.append(Paragraph(line[2:], styles["Title"]))
            elif line.startswith("## "):
                story.append(Paragraph(line[3:], styles["Heading2"]))
                story.append(Spacer(1, 0.2*cm))
            elif line.startswith("### "):
                story.append(Paragraph(line[4:], styles["Heading3"]))
            elif line.startswith("- "):
                story.append(Paragraph(line, styles["Normal"]))
            elif line.startswith("*") and line.endswith("*"):
                story.append(Paragraph(line[1:-1], styles["Italic"]))
            elif line.startswith("**") and "**" in line[2:]:
                content = line.replace("**", "")
                story.append(Paragraph(content, styles["Normal"]))
            else:
                story.append(Paragraph(line, styles["Normal"]))

        doc.build(story)
        buffer.seek(0)
        return buffer.read()

    # ── Excel 導出 ───────────────────────────────────────────────────────────

    async def generate_excel(
        self,
        user_id: int,
        report_type: ReportType,
        reference_date: Optional[datetime] = None,
    ) -> bytes:
        """
        生成 Excel 報告

        需要 openpyxl: pip install openpyxl
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            logger.error("openpyxl not installed, cannot generate Excel")
            return b"ERROR: openpyxl not installed"

        data = await self.collect_report_data(user_id, report_type, reference_date)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 移除預設 sheet

        # Sheet 1: Summary
        ws1 = wb.create_sheet("摘要")
        ws1.append(["黃金分析報告"])
        ws1.append(["報告類型", data.report_type])
        ws1.append(["報告標題", data.title])
        ws1.append(["報告週期", f"{data.period_start.strftime('%Y-%m-%d')} ~ {data.period_end.strftime('%Y-%m-%d')}"])
        ws1.append(["生成時間", data.generated_at.strftime("%Y-%m-%d %H:%M:%S")])
        ws1.append([""])

        # 決策統計
        ds = data.decisions
        ws1.append(["決策統計"])
        ws1.append(["總決策數", ds.get("total", 0)])
        for t, cnt in ds.get("by_type", {}).items():
            ws1.append([t, cnt])
        ws1.append([""])

        # 告警統計
        ws1.append(["告警統計"])
        ws1.append(["總告警數", data.alerts.get("total", 0)])
        ws1.append(["已觸發", data.alerts.get("triggered", 0)])
        ws1.append(["活躍", data.alerts.get("active", 0)])
        ws1.append([""])

        # 組合績效
        if data.portfolio_performance:
            ws1.append(["組合績效"])
            for name, perf in data.portfolio_performance.items():
                ws1.append([name])
                for k, v in perf.items():
                    ws1.append([f"  {k}", v])

        # Sheet 2: Decisions Detail
        ws2 = wb.create_sheet("決策明細")
        ws2.append(["ID", "類型", "資產", "信號強度", "信心度", "是否執行", "創建時間", "理由"])
        for d in data.decisions.get("details", []):
            ws2.append([
                d.get("id"), d.get("type"), d.get("asset"),
                d.get("signal_strength"), d.get("confidence"),
                "是" if d.get("is_executed") else "否",
                d.get("created_at"), d.get("reason_zh", ""),
            ])

        # 格式化
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for ws in [ws1, ws2]:
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 30
            if ws == ws2:
                ws.column_dimensions["H"].width = 50

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    # ── 入口方法 ─────────────────────────────────────────────────────────────

    async def generate_report(
        self,
        user_id: int,
        report_type: ReportType,
        output_format: str = "markdown",
        reference_date: Optional[datetime] = None,
    ) -> bytes | str:
        """
        統一報告生成入口

        Args:
            user_id: 用戶ID
            report_type: 報告類型
            output_format: "markdown" | "pdf" | "excel"
            reference_date: 參考日期

        Returns:
            markdown: str
            pdf/excel: bytes
        """
        if output_format == "markdown":
            data = await self.collect_report_data(user_id, report_type, reference_date)
            return self.generate_markdown(data)
        elif output_format == "pdf":
            return await self.generate_pdf(user_id, report_type, reference_date)
        elif output_format == "excel":
            return await self.generate_excel(user_id, report_type, reference_date)
        else:
            raise ValueError(f"Unsupported format: {output_format}")
